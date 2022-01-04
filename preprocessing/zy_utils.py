"""
description: my dependent utils
author: zhangyan
date: 2021-04-02 17:42
"""

import os
import json
import shutil
import yaml
import pandas as pd
import openpyxl as xl
from openpyxl.styles import Font, Alignment
from pypinyin import pinyin, NORMAL
import xml.etree.ElementTree as ET
import cv2
import math
import numpy as np


IMG_TYPES = ['jpg', 'png', 'JPG', 'PNG']

def create_empty_json_instance(img_file_path: str):
    '''
    :param img_file_path: img路径
    :return: 构建一个空的labelme json instance对象
    '''
    instance = {'version': '1.0',
                'shapes': [],
                'imageData': None,
                'imagePath': img_file_path[img_file_path.rindex(os.sep)+1:]}
    img = cv2.imread(img_file_path)
    instance['imageHeight'], instance['imageWidth'], instance['imageDepth'] = img.shape
    # instance_to_json(instance, img_file_path[:img_file_path.rindex('.')]+'.json')
    return instance

def json_to_instance(json_file_path):
    with open(json_file_path, 'r', encoding='utf-8') as f:
        instance = json.load(f)
    return instance

def instance_to_json(instance, json_file_path):
    with open(json_file_path, 'w', encoding='utf-8') as f:
        content = json.dumps(instance, ensure_ascii=False, indent=2)
        f.write(content)

def yaml_to_instance(yaml_file_path):
    """
    yaml_file_path: yaml文件路径
    return yaml_instance
    """
    with open(yaml_file_path, 'r', encoding='utf-8') as f:
        config = f.read()
    cfg = yaml.load(config, yaml.FullLoader)
    return cfg

def filtrate_file(path):
    list = os.listdir(path)
    for obj in list:
        file_path = os.path.join(path, obj)
        if not os.path.isfile(file_path) or obj[obj.rindex('.') + 1:] not in ['json', 'jpg', 'png']: continue

def word_to_pinyin(word):
    """
    @param word:
    @return:
    """
    # pinyin return [[py1],[py2],...,[pyn]]
    s = ''
    for i in pinyin(word, style=NORMAL):
        s += i[0].strip()
    return s

def read_txt(path):
    with open(path, "r", encoding='utf-8') as f:  # 打开文件
        data = f.readlines()  # 读取文件
    return data

# 写入新的excel   ### 内容， 保存路径, 路径下sheet， 表头， 表头列， 插入位置行数，插入位置列数
def content_to_excel(content, save_path, sheet_name = None, header = False, index=False, row=None, col=None):
    excel_data = pd.DataFrame(content)
    writer = pd.ExcelWriter(save_path)  # 写入Excel文件
    excel_data.to_excel(writer, sheet_name=sheet_name, header=header, index=index, startrow=row, startcol=col)
    writer.save()
    writer.close()

# 追加写入excel（可指定位置）
def write_excel_xlsx_append(file_path, data_name, data, row=0, col=0, sheet_name='sheet1'):
    workbook = xl.load_workbook(file_path)  # 打开工作簿
    sheet = workbook[sheet_name]
    for i in range(0, len(data)):
        for j in range(0, len(data[i])):
            sheet.cell(row=(i + row + 1), column=(j + col + 1), value=data[i][j])  # 追加写入数据，注意是从i+row行，j + col列开始写入
    workbook.save(file_path)  # 保存工作簿
    print("xlsx格式表格【追加】写入{}成功！".format(data_name))

# 创建一个excel，sheet
def create_empty_excel(save_path, sheet):
    wb = xl.Workbook()
    ws = wb.create_sheet(sheet, 0)
    wb.save(save_path)

# 读取excel中的sheet
def read_excel(excel_path, sheet_name):
    book = xl.load_workbook(excel_path)
    sheet = book[sheet_name]
    return sheet

# 美化excel
def beautify_excel(excel_path):
    font = Font(name='宋体', size=11, color='FF000000', bold=True, italic=False)
    align = Alignment(horizontal='center', vertical='center', wrap_text=False)

    book = xl.load_workbook(excel_path)  # 加载excel
    sheet_names = book.sheetnames  # 获取所有的sheet名称
    sheet_names.remove('Sheet')

    for sheet in sheet_names:
        for row in book['{}'.format(sheet)]['A1:A5']:
            for cell in row:
                cell.font = font
                cell.alignment = align

        for row in book['{}'.format(sheet)]['A12:Z19']:
            for cell in row:
                cell.alignment = align

    book.save(excel_path)

# 创建节点
def create_Node(element, text=None):
    elem = ET.Element(element)
    elem.text = text
    return elem

# 链接节点到根节点
def link_Node(root, element, text=None):
    """
    @param root: element的父节点
    @param element: 创建的element子节点
    @param text: element节点内容
    @return: 创建的子节点
    """
    if text != None:
        text = str(text)
    element = create_Node(element, text)
    root.append(element)
    return element

def compute_iou(bbox1, bbox2):
    """
    compute iou
    """
    bbox1xmin = bbox1[0]
    bbox1ymin = bbox1[1]
    bbox1xmax = bbox1[2]
    bbox1ymax = bbox1[3]
    bbox2xmin = bbox2[0]
    bbox2ymin = bbox2[1]
    bbox2xmax = bbox2[2]
    bbox2ymax = bbox2[3]

    area1 = (bbox1ymax - bbox1ymin) * (bbox1xmax - bbox1xmin)
    area2 = (bbox2ymax - bbox2ymin) * (bbox2xmax - bbox2xmin)

    bboxxmin = max(bbox1xmin, bbox2xmin)
    bboxxmax = min(bbox1xmax, bbox2xmax)
    bboxymin = max(bbox1ymin, bbox2ymin)
    bboxymax = min(bbox1ymax, bbox2ymax)
    if bboxxmin >= bboxxmax:
        return 0
    if bboxymin >= bboxymax:
        return 0

    area = (bboxymax - bboxymin) * (bboxxmax - bboxxmin)
    iou = area / (area1 + area2 - area)
    return iou

class Box:
    # x, y是左上角坐标
    def __init__(self, x, y, w, h, category=None, confidence=None):
        self.category = category
        self.confidence = confidence
        self.x = x
        self.y = y
        self.w = w
        self.h = h

    def get_area(self):
        return self.w * self.h

    def get_iou(self, box2):
        inter_area = calculate_inter_area(self, box2)
        return inter_area/(self.get_area()+box2.get_area()-inter_area)

def calculate_inter_area(box1, box2):
    '''
    :param box1: Box对象
    :param box2: Box对象
    :return: box1与box2的交面积
    '''
    left_x, left_y = max([box1.x, box2.x]), max([box1.y, box2.y])
    right_x, right_y = min([box1.x + box1.w, box2.x + box2.w]), min([box1.y + box1.h, box2.y + box2.h])
    height = right_y - left_y
    width = right_x - left_x
    area = height * width if height>0 and width>0 else 0
    return area

# -----以下代码用来创建文件夹-----
def make_dir(base_path):
    if not os.path.exists(base_path):
        os.mkdir(base_path)

def make_dir2(base_path, folders):
    if not os.path.exists(base_path):
        os.mkdir(base_path)
    for folder in folders:
        folder_path = os.path.join(base_path, folder)
        if not os.path.exists(folder_path):
            os.mkdir(folder_path)

def make_dir3(base_path, folders, subfolders):
    if not os.path.exists(base_path):
        os.mkdir(base_path)
    for folder in folders:
        folder_path = os.path.join(base_path, folder)
        if not os.path.exists(folder_path):
            os.mkdir(folder_path)
        for subfolder in subfolders:
            subfolder_path = os.path.join(folder_path, subfolder)
            if not os.path.exists(subfolder_path):
                os.mkdir(subfolder_path)
# -----以上代码用来创建文件夹-----

# 移动文件夹下所有指定尾缀文件到另一个文件夹
def move_specify_file(input_path, file_type, output_path):
    file_list = os.listdir(input_path)
    for file in file_list:
        if file.endswith(file_type):
            shutil.move(os.path.join(input_path, file), output_path)

# -----以下代码用来进行坐标转换-----
def extract_xys(axiss):
    '''
    :param axiss: xml中的坐标系父节点
    :return: list[x1,y1,...,xn,yn]
    '''
    return [float(axis.text) for axis in axiss if len(axis.text) > 5]

def points_to_xywh(obj):
    '''
    :param obj: labelme instance中待检测目标obj{}
    :return: box左上坐标+wh
    '''
    points = obj['points']
    shape_type = obj['shape_type']
    if shape_type == 'circle':
        center = [points[0][0], points[0][1]]
        radius = math.sqrt((points[1][0]-center[0])**2+(points[1][1]-center[1])**2)
        return [center[0]-radius-1, center[1]-radius-1, 2*radius+3, 2*radius+3]
    xs = [point[0] for point in points]
    ys = [point[1] for point in points]
    min_x, max_x = min(xs), max(xs)
    min_y, max_y = min(ys), max(ys)
    return [min_x-1, min_y-1, max_x-min_x+3, max_y-min_y+3]

def points_to_center(obj):
    '''
    :param obj: labelme instance中待检测目标obj{}
    :return: box中心坐标
    '''
    points = obj['points']
    shape_type = obj['shape_type']
    if shape_type == 'circle':
        return points[0][0], points[0][1]
    xs = [point[0] for point in points]
    ys = [point[1] for point in points]
    min_x, max_x = min(xs), max(xs)
    min_y, max_y = min(ys), max(ys)
    return (min_x+max_x)/2, (min_y+max_y)/2

def yolo_to_xywh(line):
    line = line.split(' ')
    x = float(line[1]) - float(line[3]) / 2
    y = float(line[2]) - float(line[4]) / 2
    w = float(line[3])
    h = float(line[4])
    category = int(line[0])
    return [x, y, w, h, category]
# -----以上代码用来进行坐标转换-----

# 字典对齐---以dic_a为基准，dic_b向dic_a对齐，补0
def dic_align(dic_a, dic_b):
    c = [i for i in dic_a if i not in dic_b]
    for i in c:
        dic_b[i] = 0
    return dic_b

def grid_search(a, b):
    """
    @param a: type->list [start, stop, N]
    @param b: type->list [start, stop, N]
    @return: two-dimensional array
    """
    # x, y = np.meshgrid(np.linspace(a[0], a[1], a[2]), np.linspace(b[0], b[1], b[2]))
    # cartesian_arr = np.array([x.ravel(),y.ravel().T])
    # return cartesian_arr.T
    x, y = np.meshgrid(np.linspace(a[0], a[1], a[2]), np.linspace(b[0], b[1], b[2]))
    cartesian_arr = np.array([x.ravel(),y.ravel().T])
    return np.round(cartesian_arr.T, 2)


################# 以下paul ###################

def instance_clean(instance):
    '''
    :param instance: labelme json instance
    :return: 将不良points进行清洗、更正
    '''
    for obj in instance['shapes']:
        points = obj['points']
        if obj['shape_type'] in ('line', 'linestrip'):
            # 排除标注小组的重复落点
            points_checked = [points[0]]
            for point in points:
                if point != points_checked[-1]:
                    points_checked.append(point)
            obj['points'] = points_checked
            points = points_checked
            # 排除标注小组的往返落点
            if len(points) >= 3:
                temp = get_angle((points[-3][0]-points[-2][0], points[-3][1]-points[-2][1]), (points[-1][0]-points[-2][0], points[-1][1]-points[-2][1]))
                if temp > -0.001 and temp < 0.001: del points[-1]
            if len(points) == 1: obj['shape_type'] = 'point'
            elif len(points) == 2: obj['shape_type'] = 'line'
            else: obj['shape_type'] = 'linestrip'

def instance_points_to_polygon(instance):
    '''
    :param instance: labelme json instance
    :return: 将instance['shapes']中points的标签，由rectangle和circle变为polygon，从而更好地进行crop和fill
    '''
    objs = instance['shapes']
    for obj in objs:
        shape_type = obj['shape_type']
        points = obj['points']
        if shape_type == 'rectangle':
            xs = [point[0] for point in points]
            ys = [point[1] for point in points]
            min_x, min_y = min(xs), min(ys)
            max_x, max_y = max(xs), max(ys)
            obj['points'] = [[min_x, min_y], [max_x, min_y], [max_x, max_y], [min_x, max_y]]
            obj['shape_type'] = 'polygon'
        elif shape_type == 'circle':
            center = [points[0][0], points[0][1]]
            radius = math.sqrt((points[1][0]-center[0])**2+(points[1][1]-center[1])**2)
            obj['points'] = []
            obj['shape_type'] = 'polygon'
            for i in range(0, 360, 10):
                obj['points'].append([center[0]+math.cos(math.pi*i/180)*radius, center[1]+math.sin(math.pi*i/180)*radius])

# -----以下代码求两线段交点坐标-----
class Point(object):
    def __init__(self, x=0, y=0):
        self.x = x
        self.y = y

class Line(object):
    # a=0, b=0, c=0
    def __init__(self, p1, p2):
        self.p1 = p1
        self.p2 = p2

def getLinePara(line):
    line.a =line.p1.y - line.p2.y
    line.b = line.p2.x - line.p1.x
    line.c = line.p1.x *line.p2.y - line.p2.x * line.p1.y

def getCrossPoint(l1, l2):
    getLinePara(l1)
    getLinePara(l2)
    d = l1.a * l2.b - l2.a * l1.b
    p=Point()
    if d == 0: return None
    p.x = (l1.b * l2.c - l2.b * l1.c)*1.0 / d
    p.y = (l1.c * l2.a - l2.c * l1.a)*1.0 / d
    return p

def get_cross_point(x1, y1, x2, y2, x3, y3, x4, y4):
    p1 = Point(x1, y1)
    p2 = Point(x2, y2)
    l1 = Line(p1, p2)
    p3 = Point(x3, y3)
    p4 = Point(x4, y4)
    l2 = Line(p3, p4)
    cp = getCrossPoint(l1, l2)
    if cp == None \
            or cp.x < min([x1, x2])-0.01 \
            or cp.x > max([x1, x2])+0.01 \
            or cp.y < min([y1, y2])-0.01 \
            or cp.y > max([y1, y2])+0.01 \
            or cp.x < min([x3, x4])-0.01 \
            or cp.x > max([x3, x4])+0.01 \
            or cp.y < min([y3, y4])-0.01 \
            or cp.y > max([y3, y4])+0.01:
        return None
    else:
        return [cp.x, cp.y]
# -----以上求两线段交点坐标-----

def points_to_coco_segmentation(obj, line_pixel):
    '''
    :param obj: labelme instance中待检测目标obj{}
    :param line_pixel: labelme中line、linestrip points的加宽像素值
    :return: coco segmentation[[x1,y1,x2,y2,...,xn,yn]]
    '''
    points = obj['points']
    shape_type = obj['shape_type']
    if shape_type == 'rectangle':
        xs = [point[0] for point in points]
        ys = [point[1] for point in points]
        min_x, min_y = min(xs), min(ys)
        max_x, max_y = max(xs), max(ys)
        result = [[min_x, min_y, max_x, min_y, max_x, max_y, min_x, max_y]]
    elif shape_type == 'circle':
        center = [points[0][0], points[0][1]]
        radius = math.sqrt((points[1][0]-center[0])**2+(points[1][1]-center[1])**2)
        temp = []
        for i in range(0, 360, 10):
            temp.append(center[0]+math.cos(math.pi*i/180)*radius)
            temp.append(center[1]+math.sin(math.pi*i/180)*radius)
        result = [temp]
    elif shape_type == 'line' or shape_type == 'linestrip':
        result = [line_pixel_widen(points, line_pixel)]
    else:
        result = [np.asarray(points).flatten().tolist()]
    return result

def line_pixel_widen(points, line_pixel):
    '''
    :param points: labelme中标签为line、linestrip的points
    :param line_pixel: 加宽的像素点
    :return: 返回coco中直线segmentation的坐标点
    '''
    line1 = []
    line2 = []
    for i, point in enumerate(points):
        belong_to_line1 = True
        if i == 0:
            vector2 = (points[i+1][0] - point[0], points[i+1][1] - point[1])
            angle_horiz = get_horiz_angle(vector2) + math.pi/2
            line1.append(perturbation_around_point(point, angle_horiz, line_pixel)[0])
            line2.append(perturbation_around_point(point, angle_horiz, line_pixel)[1])
            continue
        elif i == len(points)-1:
            vector1 = (points[i-1][0] - point[0], points[i-1][1] - point[1])
            angle_horiz = get_horiz_angle(vector1) + math.pi/2
            perturb_points = perturbation_around_point(point, angle_horiz, line_pixel)
            if get_cross_point(perturb_points[0][0], perturb_points[0][1], line1[-1][0], line1[-1][1], points[i-1][0], points[i-1][1], point[0], point[1]) != None:
                belong_to_line1 = False
        else:
            vector1 = (points[i-1][0] - point[0], points[i-1][1] - point[1])
            vector2 = (points[i+1][0] - point[0], points[i+1][1] - point[1])
            angle_horiz = get_mid_horiz_angle(vector1, vector2)
            radius = line_pixel/math.sin(get_angle(vector1, vector2)/2)
            perturb_points = perturbation_around_point(point, angle_horiz, radius)
            if get_cross_point(perturb_points[0][0], perturb_points[0][1], line1[-1][0], line1[-1][1], points[i-1][0], points[i-1][1], point[0], point[1]) != None:
                belong_to_line1 = False
        if belong_to_line1:
            line1.append(perturb_points[0])
            line2.append(perturb_points[1])
        else:
            line1.append(perturb_points[1])
            line2.append(perturb_points[0])
    # import matplotlib.pyplot as plt
    # line = [[points[i][0], points[i][1], points[i - 1][0], points[i - 1][1]] for i in range(1, len(points))]
    # plt.plot([p[0] for p in points], [p[1] for p in points], 'g-')
    # plt.plot([l[0] for l in line1], [l[1] for l in line1], 'b:')
    # plt.plot([l[0] for l in line2], [l[1] for l in line2], 'b:')
    # plt.show()
    return np.asarray(line1 + list(reversed(line2))).flatten().tolist()

def perturbation_around_point(point, angle, radius):
    '''
    :param point: 目标坐标点
    :param angle: 微扰角度
    :param radius: 微扰幅度
    :return: 返回一个坐标点的周围两个微扰点
    '''
    return [point[0] + radius * math.cos(angle), point[1] + radius * math.sin(angle)],\
           [point[0] - radius * math.cos(angle), point[1] - radius * math.sin(angle)]

def get_mid_horiz_angle(vector1, vector2):
    '''
    :param vector1: 向量1
    :param vector2: 向量2
    :return: 两个向量的中间向量与水平线(1,0)的夹角
    '''
    angle_horiz_1 = get_horiz_angle(vector1)
    angle_horiz_2 = get_horiz_angle(vector2)
    return (angle_horiz_1 + angle_horiz_2)/2

def get_horiz_angle(vector):
    '''
    :param vector: 一个向量
    :return: 该向量和水平线(1,0)的夹角(-180-180)
    '''
    angle_horiz = get_angle((1, 0), vector) if vector[1] > 0 else -get_angle((1, 0), vector)
    return angle_horiz

def get_angle(vector1, vector2):
    '''
    :param vector1: 向量1
    :param vector2: 向量2
    :return: 向量之间的夹角(0-180)
    '''
    inner_product = vector1[0]*vector2[0] + vector1[1]*vector2[1]
    cosin = inner_product/(math.sqrt((vector1[0]**2+vector1[1]**2)*(vector2[0]**2+vector2[1]**2)))
    return math.acos(cosin)

def crop_is_empty(instance, crop_size, iou_thres=0.2):
    '''
    :param instance: labelme json instance
    :param crop_size: crop范围[上，下，左，右]
    :return: bool值
    '''
    flag = True
    for obj in instance['shapes']:
        if obj_in_crop(obj, crop_size, iou_thres):
            flag = False
            break
    return flag

def obj_in_crop(obj, crop_size, iou_thres=0.2):
    '''
    :param points: labelme json中一个obj的points
    :param crop_size: crop范围[上，下，左，右]
    :param iou_thres: iou阈值
    :return: bool值
    '''
    crop_box = Box(crop_size[2], crop_size[0], crop_size[3] - crop_size[2], crop_size[1] - crop_size[0])
    x, y, w, h = points_to_xywh(obj)
    obj_box = Box(x, y, w, h)
    inter_area = calculate_inter_area(obj_box, crop_box)
    return inter_area != 0 and inter_area/obj_box.get_area() >= iou_thres

def point_in_crop(point, crop_size):
    '''
    :param point: labelme json中一个obj的points-point[x,y]
    :param crop_size: crop范围[上，下，左，右]
    :return: bool值
    '''
    return point[0] > crop_size[2] and \
           point[0] < crop_size[3] and \
           point[1] > crop_size[0] and \
           point[1] < crop_size[1]
################# 以上paul ###################

if __name__ == '__main__':
    print(perturbation_around_point([1,1], -1.57/2, 1.414))

