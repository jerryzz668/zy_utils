# @Description:
# @Author     : zhangyan
# @Time       : 2021/4/19 8:48 下午

import random
from xml.dom import minidom

import openpyxl as xl
# ---------------------------class7----------------------start
# 实物csv和实物图可视化，给他实物图和实物csv生成对应的xml标注和标注合并图，显示的时候注意类别映射字典。调用：#csv_p = r'C:\Users\xie5817026\PycharmProjects\pythonProject1\0104\ProductGradeMaterialCheck.csv'
# img_p ='D:\work\data\microsoft\jalama\data\heduiji\merge_all",ShiwuHedui(img_p,csv_p),生成'D:\work\data\microsoft\jalama\data\heduiji\merge_all\outputs",'D:\work\data\microsoft\jalama\data\heduiji\merge_all\r_imgs"
import pandas as ps
from PIL import Image


# @Description:
# @Author     : zhangyan
# @Time       : 2021/1/14 3:54 下午
class Dic2xml(object):
    def __init__(self, dic, xml_save_path):
        self.dic2xml(dic, xml_save_path)

    def create_Node(self, element, text=None):
        elem = ET.Element(element)
        elem.text = text
        return elem

    def link_Node(self, root, element, text=None):
        """
        @param root: element的父节点
        @param element: 创建的element子节点
        @param text: element节点内容
        @return: 创建的子节点
        """
        if text != None:
            text = str(text)
        element = self.create_Node(element, text)
        root.append(element)
        return element

    # 保存为XML文件（美化后）
    def saveXML(self, root, filename, indent="\t", newl="\n", encoding="utf-8"):
        rawText = ET.tostring(root)
        dom = minidom.parseString(rawText)
        with open(filename, 'w', encoding="utf-8") as f:
            dom.writexml(f, "", indent, newl, encoding)

    def get_dic_data(self, key, value):
        save_name = key.split('.')[0] + '.xml'
        anno = value.get('anno')
        w = value.get('w')
        h = value.get('h')
        return save_name, save_name, anno, None, 'true', w, h, 3

    def generate_xml(self, key, value, xml_save_path):
        save_name, xmlpath, anno, time_label, image_label, width, height, depth = self.get_dic_data(key, value)

        root = ET.Element("doc")  # 创建根结点

        path = self.link_Node(root, 'path', xmlpath)  # 创建path节点
        outputs = self.link_Node(root, 'outputs')
        object = self.link_Node(outputs, 'object')

        for i in range(len(anno)):
            item = self.link_Node(object, 'item')  # 创建item节点

            label = anno[i][4]  # 获取label
            width_points_line = 2  # 点或线的width
            shape_type = 'rectangle'

            name = self.link_Node(item, 'name', label)  # 添加json信息到item中
            width_2 = self.link_Node(item, 'width', width_points_line)

            if shape_type == 'rectangle':
                bndbox = self.link_Node(item, 'bndbox')
                xmin = self.link_Node(bndbox, 'xmin', int(anno[i][0]))
                ymin = self.link_Node(bndbox, 'ymin', int(anno[i][1]))
                xmax = self.link_Node(bndbox, 'xmax', int(anno[i][2]))
                ymax = self.link_Node(bndbox, 'ymax', int(anno[i][3]))

            status = self.link_Node(item, 'status', str(1))

        time_labeled = self.link_Node(root, 'time_labeled', time_label)  # 创建time_labeled节点
        labeled = self.link_Node(root, 'labeled', image_label)
        size = self.link_Node(root, 'size')
        width = self.link_Node(size, 'width', width)
        height = self.link_Node(size, 'height', height)
        depth = self.link_Node(size, 'depth', depth)

        save_path = xml_save_path  # os.path.join(xml_save_path, save_name)
        # if not os.path.exists(xml_save_path):
        #     os.makedirs(xml_save_path)
        # 保存xml文件
        self.saveXML(root, save_path)
        print('{}'.format(save_name) + ' has been transformed!')

    def dic2xml(self, dic, xml_save_path):
        t = time.time()
        for key, value in dic.items():
            self.generate_xml(key, value, xml_save_path)
        print(time.time() - t)


# if __name__ == '__main__':
#     dic = {'123.jpg': {'anno': [(132, 243, 355, 467, '刮伤'), (51, 61, 72, 82, '异色')], 'w': 512, 'h': 512},
#            '456.jpg': {'anno':[(11, 21, 31, 41, '擦伤'), (11, 22, 33, 41, '白点')], 'w': 512, 'h': 512}}
#
#     xml_save_path = r'C:\Users\xie5817026\PycharmProjects\pythonProject1\0104\xml'
#     Dic2xml(dic, xml_save_path)
class ShiwuHedui(object):
    def __init__(self, source_p, csv_p, xml_p):

        # 图像位置
        dic_wh = self.w_h(source_p)  # 图像存放位置#获取所有图像的wh字典
        if not os.path.exists(xml_p):
            os.makedirs(xml_p)
        for imgs_name in os.listdir(source_p):
            if imgs_name.endswith('jpg'):
                csv_name = imgs_name.replace('jpg', 'csv')
                xml_name = imgs_name.replace('jpg', 'xml')
                csv_pp = os.path.join(csv_p, csv_name)
                xml_save_path = os.path.join(xml_p, xml_name)
                img_dic_c = self.read_csv(csv_pp, dic_wh, imgs_name)  # csv格式 #读取一张图像csv
                print('img_dic_c', img_dic_c)
                Dic2xml(img_dic_c, xml_save_path)

    def w_h(self, p):
        dic_wh = {}
        print('---')
        for i in os.listdir(p):
            i_p = os.path.join(p, i)
            if not os.path.isfile(i_p) or i[i.rindex('.') + 1:] not in ['json', 'jpg', 'png']: continue
            i_p = os.path.join(p, i)
            # data = cv2.imread(i_p)
            data = Image.open(i_p)
            dic_wh[i] = data.size
            print(data.size)
        return dic_wh

    def read_csv(self, csv_path, dic_wh, img_name):
        r = ps.read_csv(csv_path,
                        usecols=['photo_id', 'product_id', 'channel_id', 'class_name', 'xmin', 'ymin', 'bb_width',
                                 'bb_height'])
        # r = ps.read_csv(csv_path,usecols=['任务号','工件号','图号','缺陷','PointX','PointY','Width','Height'])
        imgs_dic = {}
        x_min = r['xmin']
        y_min = r['ymin']
        w = r['bb_width']
        h = r['bb_height']
        label = r['class_name']

        for i in range(len(label)):
            img_name = img_name  # '{}-{}-{}.jpg'.format(task_ids[i],gongjian_ids[i],img_ids[i])
            x_max = x_min[i] + w[i]
            y_max = y_min[i] + h[i]
            if img_name in imgs_dic:
                imgs_dic[img_name].append((x_min[i], y_min[i], x_max, y_max, label[i]))
            else:
                imgs_dic[img_name] = [(x_min[i], y_min[i], x_max, y_max, label[i])]
        img_dic_c = {}
        for i in imgs_dic:
            try:
                w, h = dic_wh[i]
                one_img_dic = {}
                one_img_dic['anno'] = imgs_dic[i]
                one_img_dic['w'] = w
                one_img_dic['h'] = h
                img_dic_c[i] = one_img_dic
            except:
                print('--')
        print(img_dic_c)
        return img_dic_c


# ---------------------------class7----------------------end

from xml.etree import ElementTree as ET
import multiprocessing
import pypinyin
import cv2

'''
@Xml2Labelme:
            主要功能：将mark工具标注的xml标注转为json标注，方便labelme查看及生成coco格式数据
            关键处理：1.汉字转拼音 pypinyin工具包；2.检测标注类别错标并调整类别，modify_type函数；
                    3.xml结构转为coco结构，transform_xml_2labelme函数
@author: lijianqing
@date: 2020/11/11 13:19
'''


class Xml2Labelme(object):
    def __init__(self, annotation_root_path, jsons_path, result_write, process_num=8):
        self.xml_outputs_path = os.path.join(annotation_root_path, 'outputs')
        if not os.path.exists(jsons_path):
            os.makedirs(jsons_path)
        self.jsons_path = jsons_path
        self.result_write = result_write
        s = time.time()
        # self.main(self.xml_outputs_path,self.jsons_path,process_num)
        self.cut_label = self.main_share(self.xml_outputs_path, self.jsons_path, process_num)
        print('xml2labelme run time:', time.time() - s)

    def pinyin(self, word):  # zhangsan = pinyin('张山')
        s = ''
        for i in pypinyin.pinyin(word, style=pypinyin.NORMAL):
            s += ''.join(i)
        return s

    def modify_type(self, type, points_l):
        type_index = {0: 'none', 1: 'point', 2: 'line', 3: 'polygon', 4: 'bndbox', 5: 'ellipse'}
        if len(points_l) < 3:  # 长度为0，1,2的
            if type != 'ellipse' and type != 'bndbox':  # 类别不为圆和bndbox的情况下按长度设置类别，长度为0的都为none,长度为1的都是点，长度为2的都是线，
                type = type_index[len(points_l)]
            elif len(points_l) != 2:
                type = type_index[len(points_l)]  # 类别为圆或bndbox，但长度小于2的，按长度设置类别，长度为0的都为none,长度为1的都是点，
            else:  # 类别为圆或bndbox，长度为2的，符合圆或bndbox的标准，不做修改
                type = type
        elif len(points_l) < 4:  # 长度为3的
            if type != 'line':  # 类别不为line的按多边形处理
                type = type_index[len(points_l)]
            else:  # 类别为line的，符合line标准，不做修改
                type = type
        else:  # 长度大于或等于4的
            if type != 'line' and type != 'polygon':  # 类别既不是line,也不是多边形的，按多边形polygon处理
                type = type_index[3]
            else:  # 类别为线，或多边形的，不做修改
                type = type
        return type

    def transform_xml_2labelme(self, xml_path, save_json, image_name, class_py_ch_dic):
        transform_shapes_dic = {'polygon': 'polygon', 'line': 'linestrip', 'bndbox': 'rectangle', 'point': 'circle',
                                'ellipse': 'circle'}
        tree = ET.parse(xml_path)
        root_node = tree.getroot()
        imagePath = '{}.jpg'.format(image_name)
        xmlPath = root_node[0].text
        print('class_py_ch_dic--', xmlPath)
        imageWidth = int(root_node[4][0].text)  # json_obj['imageWidth']=0
        imageHeight = int(root_node[4][1].text)  # json_obj['imageHeight']=0
        imageDepth = int(root_node[4][2].text)  # null
        imageLabeled = root_node[3].text
        time_Labeled = root_node[2].text
        shapes = []
        print('root_node[1][0]', root_node[1][0])
        for i in root_node[1][0]:
            dic_instance = {}
            dic_instance['label'] = self.pinyin(i[0].text)
            class_py_ch_dic[self.pinyin(i[0].text)] = i[0].text
            points_l = []
            point_axis = []
            flag = 0
            # print('kd----------:',i[1].text,'--',i[2].tag)#像素宽度
            for j in i[2].iter():
                if flag != 0:
                    # print('j.text,shape:',j.text,i[2].tag,i[1].text,)
                    point_axis.append(float(j.text))
                    if len(point_axis) == 2:
                        points_l.append(point_axis)
                        point_axis = []
                else:
                    flag = 1
            xml_tag = i[2].tag
            # modify type xml
            i_2_tag = self.modify_type(xml_tag, points_l)
            # point2circle
            if i_2_tag == 'point' and len(points_l) == 1:
                r = int(i[1].text) / 2
                x, y = points_l[0]
                r_x = x + r
                r_y = y + r
                points_l.append([r_x, r_y])
                i_2_tag = 'ellipse'

            # 2 labelme json after modify type in xml
            try:
                dic_instance['shape_type'] = transform_shapes_dic[i_2_tag]
            except:
                print('标注类别为空或错误，请检测。', i[2].tag, xml_path)
            dic_instance['width'] = i[1].text
            dic_instance['points'] = points_l
            dic_instance['group_id'] = ''
            dic_instance['flags'] = {}  # i[3].text#state
            try:
                dic_instance['level'] = i[4].text  # 'level'
            except:
                # print('无照片严重程度')
                a = 0
            try:
                dic_instance['mlevel'] = i[5].text  # 'mlevel'
            except:
                # print('无缺陷严重程度')
                a = 0
            try:
                dic_instance['describe'] = i[6].text  # 'describe'
            except:
                # print('无描述')
                a = 0
            if len(points_l) > 0:  # 标注长度大于0时为可用的有效标注
                shapes.append(dic_instance)
        dic_all = {}
        dic_all['version'] = '1.0'
        dic_all['flags'] = {}
        dic_all['shapes'] = shapes
        dic_all['imagePath'] = imagePath
        dic_all['xmlPath'] = xmlPath
        # print('正在处理：',imagePath)
        dic_all['imageData'] = None
        dic_all['imageHeight'] = imageHeight
        dic_all['imageWidth'] = imageWidth
        dic_all['imageDepth'] = imageDepth
        dic_all['imageLabeled'] = imageLabeled
        dic_all['time_Labeled'] = time_Labeled
        with open(save_json, "w", encoding="utf-8") as f:
            content = json.dumps(dic_all, ensure_ascii=False)
            f.write(content)
            f.close()

    def main_share(self, local_path, save_path, process_num=4):
        pool = multiprocessing.Pool(processes=process_num)  # 创建进程个数
        class_py_ch_dic = multiprocessing.Manager().dict()
        for i in os.listdir(local_path):
            xml_path = os.path.join(local_path, i)  # xml_path
            image_name = i.split('.xml')[0]
            json_path = os.path.join(save_path, '{}.json'.format(image_name))
            print('json_path', json_path)
            pool.apply_async(self.transform_xml_2labelme, args=(xml_path, json_path, image_name, class_py_ch_dic))
        pool.close()
        pool.join()
        print('class_py_ch_dic:', class_py_ch_dic)
        class_py_ch_dic_tuple = sorted(class_py_ch_dic.items())
        class_pinyin_list, class_ch_list = zip(*class_py_ch_dic_tuple)
        print('{}'.format(class_pinyin_list))
        class_mapers = {}
        for i, key in enumerate(class_pinyin_list):
            class_mapers[key] = i
        with open(self.result_write, 'a+', encoding='utf-8') as f:
            f.write('拼音-汉字-字典：{}\n'.format(class_py_ch_dic))
            f.write('拼音-index：{}\n'.format(class_mapers))
            f.write('拼音：{}\n'.format(class_pinyin_list))
            f.write('汉字：{}\n'.format(class_ch_list))
            f.write('类别数量：{}\n'.format(len(class_ch_list)))
        return class_pinyin_list


import pandas as pd
import itertools
import os

# import seaborn as sns
import matplotlib.pyplot as plt
from sklearn.metrics import confusion_matrix
import time
import shutil


class AnnalyResult(object):
    def __init__(self, yt_labelme, test_labelme, out_path, title_png):
        self.yt_labelme = yt_labelme
        self.test_labelme = test_labelme
        self.out_path = out_path
        self.title_png = title_png
        self.gt_class = []
        self.pre_class = []
        start_time = time.time()
        self.main()
        end_time = time.time()
        print('run time:', end_time - start_time)
        self.cm, self.gt_cate = self.compute_confmx()

    def getcm(self):
        return self.cm, self.gt_cate

    def get_points_box(self, points, type='polygon', width=2):
        points = np.array(points)
        if type == 'point' and len(points) == 1:
            box = [points[0][0] - width / 2, points[0][1] - width / 2, points[0][0] + width / 2,
                   points[0][1] + width / 2]
            return box
        if type == 'circle' and len(points) == 2:
            r = np.sqrt((points[0][0] - points[1][0]) ** 2 + (points[0][1] - points[1][1]) ** 2)
            box = [points[0][0] - r, points[0][1] - r, points[0][0] + r, points[0][1] + r]
            return box
        box = [min(points[:, 0]), min(points[:, 1]), max(points[:, 0]), max(points[:, 1])]
        return box

    def parse_para_re(self, input_json):
        with open(input_json, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data

    def save_json(self, dic, path):
        json.dump(dic, open(path, 'w', encoding='utf-8'), indent=4)
        return 0

    def compute_iou(self, bbox1, bbox2):
        """
        compute iou
        :param bbox1:
        :param bbox2:
        :return: iou
        """
        bbox1xmin = bbox1[0]
        bbox1ymin = bbox1[1]
        bbox1xmax = bbox1[2]
        bbox1ymax = bbox1[3]
        bbox2xmin = bbox2[0]
        bbox2ymin = bbox2[1]
        bbox2xmax = bbox2[2]
        bbox2ymax = bbox2[3]
        area1 = (bbox1ymax - bbox1ymin) * (bbox1xmax - bbox1xmin)
        area2 = (bbox2ymax - bbox2ymin) * (bbox2xmax - bbox2xmin)
        bboxxmin = max(bbox1xmin, bbox2xmin)
        bboxxmax = min(bbox1xmax, bbox2xmax)
        bboxymin = max(bbox1ymin, bbox2ymin)
        bboxymax = min(bbox1ymax, bbox2ymax)
        if bboxxmin >= bboxxmax:
            return 0
        if bboxymin >= bboxymax:
            return 0
        area = (bboxymax - bboxymin) * (bboxxmax - bboxxmin)
        iou = area / (area1 + area2 - area)
        return iou

    def l_g_ls(self, require, part, iou_thr, f_l_flag):  # 过检记录：某个预测框与所有gt的iou等于0，表明该预测框为过检框,一张图预测不出结果时需要考虑漏检如何计算。
        result_l = []
        for j in require:
            result_flag = 0
            re_flag = True
            gt_points = j['points']
            for k in part:
                pre_points = k['points']
                bbox_gt = self.get_points_box(gt_points, j['shape_type'])
                bbox_re = self.get_points_box(pre_points, k['shape_type'])
                iou = self.compute_iou(bbox_gt, bbox_re)

                if iou > iou_thr:
                    re_flag = False
                    break
                # if iou<iou_thr:  # <iou_thr:
                #     # print('iou:',iou),#过检和漏检与gt的iou都为0
                #     result_flag+=1

            # if result_flag==len(part):#iou为0的数量与所有预测标注的数量是否相等，若相等表明缺陷漏检，若为0的记录小于0则表明缺陷未漏检。
            if re_flag:
                print('result_lll')
                if f_l_flag == 'loujian':  # loushi
                    self.gt_class.append(j['label'])
                    self.pre_class.append('z_lou_or_guo')
                else:  # guojian
                    self.gt_class.append('z_lou_or_guo')
                    self.pre_class.append(j['label'])
                result_l.append(j)
                print('result_l', result_l)
        return result_l

    def jiandui_ls(self, require, part, iou_thr):
        jd = []
        for j in require:
            gt_points = j['points']
            for k in part:
                pre_points = k['points']
                bbox_gt = self.get_points_box(gt_points, j['shape_type'])
                bbox_re = self.get_points_box(pre_points, k['shape_type'])
                iou = self.compute_iou(bbox_gt, bbox_re)
                if iou >= iou_thr:
                    if not k in jd:
                        jd.append(k)
                        self.gt_class.append(j['label'])
                        self.pre_class.append(k['label'])
        return jd

    def compute_confmx(self):
        classes = sorted(list(set(self.gt_class)), reverse=False)  # 类别排序
        cm = confusion_matrix(self.gt_class, self.pre_class, classes)  # 根据类别生成矩阵，此处不需要转置
        cm_pro = (cm.T / np.sum(cm, 1)).T
        # print('cm',cm)
        # print('cmp',cm_pro)
        #
        self.plot_confusion_matrix(cm, classes, 'nums')
        self.plot_confusion_matrix(cm_pro, classes, 'pro', normalize=True)
        return cm, classes
        # print('confx',cm)

    def new_json(self, cz, shapes, save_json):

        new_json_dic = {}
        new_json_dic['flags'] = cz['flags']
        new_json_dic['imageData'] = cz['imageData']
        # print(cz['imageDepth'])
        # if cz.get('imageDepth') != None:
        new_json_dic['imageDepth'] = 3  # cz['imageDepth']
        new_json_dic['imageLabeled'] = ''  # cz['imageLabeled']
        new_json_dic['imagePath'] = cz['imagePath']
        new_json_dic['imageHeight'] = cz['imageHeight']
        new_json_dic['imageWidth'] = cz['imageWidth']
        new_json_dic['shapes'] = shapes
        new_json_dic['time_Labeled'] = ''  # cz['time_Labeled']
        new_json_dic['version'] = ''  # cz['version']
        print('shapes: ', len(shapes))
        if len(shapes) != 0:
            self.save_json(new_json_dic, save_json)

    def proce_compute(self, input_json, pre_json, save_path):
        gt_anno_data = self.parse_para_re(input_json)
        print('gt_json', input_json)
        pre_anno_data = self.parse_para_re(pre_json)
        gt_shapes = gt_anno_data['shapes']
        pre_shapes = pre_anno_data['shapes']
        jiandui_shapes = []
        jiandui_shapes = self.jiandui_ls(gt_shapes, pre_shapes, 0.01)
        guojian_shapes = []
        guojian_shapes = self.l_g_ls(pre_shapes, gt_shapes, 0.01, 'guojian')
        merge_gt_pre_shapes = []
        merge_gt_pre_shapes.extend(gt_shapes)
        merge_gt_pre_shapes.extend(guojian_shapes)
        loujian_shapes = []
        try:
            loujian_shapes = self.l_g_ls(gt_shapes, pre_shapes, 0.01, 'loujian')
            print('try_loujian:', loujian_shapes)
        except:
            loujian_shapes.extend(gt_shapes)
            print('except_loujian:', loujian_shapes)
        print('---', len(guojian_shapes), len(loujian_shapes), len(jiandui_shapes), len(gt_shapes),
              len(merge_gt_pre_shapes))
        guojian_path = os.path.join(save_path, 'guojian')
        loujian_path = os.path.join(save_path, 'loujian')
        jiandui_path = os.path.join(save_path, 'jiandui')
        merge_gt_pre_path = os.path.join(save_path, 'merge_gt_pre')
        if not os.path.exists(guojian_path):
            os.makedirs(guojian_path)
        if not os.path.exists(loujian_path):
            os.makedirs(loujian_path)
        if not os.path.exists(jiandui_path):
            os.makedirs(jiandui_path)
        if not os.path.exists(merge_gt_pre_path):
            os.makedirs(merge_gt_pre_path)
        img_name = gt_anno_data['imagePath']
        json_name = img_name.replace('.jpg', '.json')
        guojian_json = os.path.join(guojian_path, json_name)
        loujian_json = os.path.join(loujian_path, json_name)
        jiandui_json = os.path.join(jiandui_path, json_name)
        merge_gt_pre_json = os.path.join(merge_gt_pre_path, json_name)
        print('=====', len(guojian_shapes), len(loujian_shapes), len(jiandui_shapes), len(gt_shapes),
              len(merge_gt_pre_shapes))
        self.new_json(gt_anno_data, guojian_shapes, guojian_json)
        self.new_json(gt_anno_data, loujian_shapes, loujian_json)
        self.new_json(gt_anno_data, jiandui_shapes, jiandui_json)
        self.new_json(gt_anno_data, merge_gt_pre_shapes, merge_gt_pre_json)

    def main(self):
        for i in os.listdir(self.yt_labelme):
            if i.endswith('.json'):
                input_json = os.path.join(self.yt_labelme, i)
                pre_json = os.path.join(self.test_labelme, i)
                except_json = os.path.join(self.out_path, i)
                try:
                    self.proce_compute(input_json, pre_json, self.out_path)
                except:
                    shutil.copy(input_json, except_json)
                    print('未预测数据', input_json)

    def plot_confusion_matrix(self, cm, classes, title, normalize=False, cmap=plt.cm.Blues):
        # plt.figure()

        plt.figure(figsize=(12, 8), dpi=120)
        plt.imshow(cm, interpolation='nearest', cmap=cmap)
        plt.title('{}_{}'.format(self.title_png, title))
        plt.colorbar()
        tick_marks = np.arange(len(classes))
        plt.xticks(tick_marks, classes, rotation=90)
        plt.yticks(tick_marks, classes)

        # plt.axis("equal")
        ax = plt.gca()
        left, right = plt.xlim()
        ax.spines['left'].set_position(('data', left))
        ax.spines['right'].set_position(('data', right))
        for edge_i in ['top', 'bottom', 'right', 'left']:
            ax.spines[edge_i].set_edgecolor("white")

        thresh = cm.max() / 2.
        for i, j in itertools.product(range(cm.shape[0]), range(cm.shape[1])):
            num = float('{:.2f}'.format(cm[i, j])) if normalize else int(cm[i, j])
            plt.text(j, i, num,
                     verticalalignment='center',
                     horizontalalignment="center",
                     color="white" if num > thresh else "black")
        plt.ylabel('ground turth')
        plt.xlabel('predict')
        plt.tight_layout()
        save_p = os.path.join(self.out_path, './{}_{}.png'.format(self.title_png, title))
        cm_txt = save_p.replace('.png', '.txt')
        with open(cm_txt, 'a+') as f:
            f.write('{}:\n'.format(title))
            f.write(str(cm))
            f.write('\n')
        # plt.savefig(save_p, transparent=True, dpi=800)
        plt.savefig(save_p, transparent=True, dpi=300)
        # plt.show()


# confusion_mtx_to_report
def confusion_mtx_to_report(data):
    """
    @param data: confusion_mtx
    @return: gt_num, loushi, loujian_ratio, jianchu, guojian, model_guojian_ratio, xianchang_guojian_ratio
    """
    loushi = []
    gt_num = []
    loujian_ratio = []
    guojian = []
    jianchu = []
    model_guojian_ratio = []
    xianchang_guojian_ratio = []

    # 漏检数量
    for i in range(len(data)):
        loushi.append(data[i][-1])
    loushi.pop()

    # gt数量
    for i in range(len(data)):
        gt_num.append(sum(data[i]))
    gt_num.pop()

    # 漏检率
    for i in range(len(loushi)):
        loujian_ratio.append(round(loushi[i] / gt_num[i], 2))
    print(loujian_ratio)

    # 每个类别过检数量
    guojian = data[-1]
    guojian = guojian[0:-1]

    # 每个类别检出总量
    jianchu = np.sum(data, axis=0)
    jianchu = jianchu[0:-1]

    # 模型过检率
    for i in range(len(jianchu)):
        model_guojian_ratio.append(round(guojian[i] / jianchu[i], 2))

    # 现场过检率
    for i in range(len(guojian)):
        xianchang_guojian_ratio.append(round(guojian[i] / gt_num[i], 2))
    print(xianchang_guojian_ratio)

    excel_content = []
    excel_content.extend(
        (gt_num, loushi, loujian_ratio, jianchu, guojian, model_guojian_ratio, xianchang_guojian_ratio))
    return excel_content


# 写入内容， 保存路径, 路径下sheet， 表头， 表头列， 插入位置行数，插入位置列数
def content_to_excel(content, save_path, sheet_name=None, header=False, index=False, row=None, col=None):
    excel_data = pd.DataFrame(content)
    writer = pd.ExcelWriter(save_path)  # 写入Excel文件
    excel_data.to_excel(writer, sheet_name=sheet_name, header=header, index=index, startrow=row, startcol=col)
    writer.save()
    writer.close()


def write_excel_xlsx_append(file_path, data_name, data, row=0, col=0, sheet_name='sheet1'):
    workbook = xl.load_workbook(file_path)  # 打开工作簿
    sheet = workbook[sheet_name]
    for i in range(0, len(data)):
        for j in range(0, len(data[i])):
            sheet.cell(row=(i + row + 1), column=(j + col + 1), value=data[i][j])  # 追加写入数据，注意是从i+row行，j + col列开始写入
    workbook.save(file_path)  # 保存工作簿
    print("xlsx格式表格【追加】写入{}成功！".format(data_name))


import json
import numpy as np
from openpyxl.drawing.image import Image as XLImage


def json_to_instance(json_file_path):
    '''
    :param json_file_path: json文件路径
    :return: json instance
    '''
    with open(json_file_path, 'r', encoding='utf-8') as f:
        instance = json.load(f)
    return instance


def cut_images(outputs_path, imgs_path, max_number=-1):
    '''
    @param max_number: 最大切图数量
    @return: NULL
    '''
    cut_images_helper(outputs_path, imgs_path, max_number, "loujian")
    cut_images_helper(outputs_path, imgs_path, max_number, "guojian")
    print("\n图片裁剪成功！")


def cut_images_helper(outputs_path, imgs_path, max_number, special_name):
    cuts_dir = os.path.join(outputs_path, special_name + "_cuts")
    if not os.path.exists(cuts_dir):
        os.mkdir(cuts_dir)
    result_dir = os.path.join(outputs_path, special_name)
    count = 0
    for f in os.listdir(result_dir):
        if f.find('.json'):
            json_content = json_to_instance(os.path.join(result_dir, f))
            shapes = json_content['shapes']
            for i in range(len(shapes)):
                # 实现上限
                if max_number != -1:
                    if count < max_number:
                        count = count + 1
                        print(count)
                    else:
                        break
                # 获取坐标
                shape = shapes[i]
                a = np.array(shape['points'])
                min_w, min_h = a.min(0)
                max_w, max_h = a.max(0)
                # 为显示清楚，padding20像素
                print(os.path.join(imgs_path, f[:-5] + '.jpg'))
                img = cv2.imread(os.path.join(imgs_path, f[:-5] + '.jpg'))
                h, w = img.shape[0], img.shape[1]
                minw = int(np.floor(min_w - 20 if (min_w - 20) > 0 else 0))
                maxw = int(np.ceil(max_w + 20 if (max_w + 20) < w else w))
                minh = int(np.floor(min_h - 20 if (min_h - 20) > 0 else 0))
                maxh = int(np.ceil(max_h + 20 if (max_h + 20) < h else h))
                # 裁剪图片
                cropped = img[minh:maxh, minw:maxw]
                cropped_name = os.path.join(cuts_dir, f[:-5] + '-' + str(i + 1) + '-' + shape['label'] + '.jpg')
                print(cropped_name)
                cv2.imwrite(cropped_name, cropped)


def add_images_to_excel(outputs_path, excel_save_path, sheet_name, position_array, ratio_array):
    '''
    @param sheet_name: 插入图片的sheet名
    @param position_array: 插入图片的起始位置 [行数, 列数]
    @param ratio_array: 插入[漏检, 过检]图片数量的百分比
    @return: NULL
    '''
    ALPHABET = [chr(i) for i in range(65, 91)]
    COL = ALPHABET
    for CHR in ALPHABET:
        CHR_LIST = list(CHR) * 26
        EXT = [i + j for i, j in zip(CHR_LIST, ALPHABET)]
        COL = COL + EXT
    book = xl.load_workbook(excel_save_path)
    sheet = book[sheet_name]
    # new_sheet = book.create_sheet(title='loujian')

    # 查找漏检图片 outputs_path/loujian_cuts
    cuts_dir = os.path.join(outputs_path, 'loujian_cuts')
    img_names = list(os.listdir(cuts_dir))
    # 决定展示图片数量
    number = 0
    if img_names:
        number = int(len(img_names) * ratio_array[0]) + 1
        show_img_names = random.sample(img_names, number)
    # 插入漏检图片
    row = position_array[0]
    col_idx = position_array[1] - 1
    sheet[COL[col_idx] + str(row)] = '漏检图片:'  # COL[col_idx] + str(row) 是 A1, B2, C1...
    if not number:
        sheet[COL[col_idx] + str(row + 1)] = '无漏检图片'
    else:
        for name in show_img_names:
            sheet[COL[col_idx] + str(row + 1)] = name
            img_name = os.path.join(cuts_dir, name)
            img = XLImage(img_name)
            # 根据图片宽度自适应排布
            w, h = img.width, img.height
            gap = 2 if w < 140 else 3
            fixed_width = 120 if w < 140 else 180  # 限定最大宽度
            img.width = fixed_width
            img.height = img.width * h / w
            sheet.add_image(img, (COL[col_idx] + str(row + 2)))
            col_idx = col_idx + gap

    # 查找过检图片 outputs_path/guojian_cuts
    cuts_dir = os.path.join(outputs_path, 'guojian_cuts')
    img_names = list(os.listdir(cuts_dir))
    # 决定展示图片数量
    number = 0
    if img_names:
        number = int(len(img_names) * ratio_array[1]) + 1
        show_img_names = random.sample(img_names, number)
    # 插入过检图片
    if not col_idx:
        col_idx = col_idx + 7
    sheet[COL[col_idx] + str(row)] = '过检图片:'
    if not number:
        sheet[COL[col_idx] + str(row + 1)] = '无过检图片'
    else:
        for name in show_img_names:
            sheet[COL[col_idx] + str(row + 1)] = name
            img_name = os.path.join(cuts_dir, name)
            img = XLImage(img_name)
            # 根据图片宽度自适应排布
            w, h = img.width, img.height
            gap = 2 if w < 140 else 3
            fixed_width = 120 if w < 140 else 180  # 限定最大宽度
            img.width = fixed_width
            img.height = img.width * h / w
            sheet.add_image(img, (COL[col_idx] + str(row + 2)))
            col_idx = col_idx + gap

    book.save(excel_save_path)
    print("\n图片插入excel成功！")


def test_to_reports(sub_file, save_path, sheet):
    table_header = [['产品代号', 'C件'],
                    ['光学面', sheet],
                    ['模型版本号', 'htc_20210419'],
                    ['测试集版本号', 'ceshi_20210419'],
                    ['测试通过条件', '刮伤漏检率<5%'],
                    ['', '黑点漏检率<2%'],
                    ['', '异物漏检率<1%'],
                    ['', ''],
                    ['', ''],
                    ['测试结果:', ''],
                    ['', '']]

    header_zhibiao = [['指标'],
                      ['gt数量'],
                      ['漏检数'],
                      ['漏检率'],
                      ['检出数'],
                      ['过检数'],
                      ['模型过检率'],
                      ['现场过检率']]

    imgs_path = os.path.join(sub_file, 'img')  # 测试img路径
    csv_path = os.path.join(sub_file, 'csv')  # csv路径   和img分开存放
    gt_json = os.path.join(sub_file, 'gt')  # biaozhu jsons

    split_result_file = os.path.join(os.path.dirname(imgs_path), 'outputs_path')  # split result file
    xml_path = os.path.join(imgs_path, 'outputs')  # 自动生成 测试结果生成的xml
    json_path = os.path.join(imgs_path, 'jsons')  # 自动生成 xml转成的json
    ShiwuHedui(imgs_path, csv_path, xml_path)  # csv_to_xml
    xml2json = Xml2Labelme(imgs_path, json_path, 'result', 8)

    print('分析标注结果生成混淆矩阵')
    annalyresult = AnnalyResult(gt_json,
                                json_path,  # pre jsons
                                split_result_file,
                                '140model_0420testdata')  # 混淆矩阵图像名字，不带后缀
    cm, gt_cate = annalyresult.getcm()
    print('gt_cate:', gt_cate)
    content = confusion_mtx_to_report(cm)

    write_excel_xlsx_append(save_path, 'table_header', table_header, 0, 0, sheet_name=sheet)  # 写入表头
    write_excel_xlsx_append(save_path, 'zhibiao_header', header_zhibiao, 11, 0, sheet_name=sheet)  # 写入指标
    write_excel_xlsx_append(save_path, 'biaoqian_header', [gt_cate[0:-1]], 11, 1, sheet_name=sheet)  # 写入标签
    write_excel_xlsx_append(save_path, 'content', content, 12, 1, sheet_name=sheet)  # 写入content

    cut_images(split_result_file, imgs_path, 100)
    add_images_to_excel(split_result_file, save_path, sheet, [22, 1], [0.1, 0.04])  # 插图：漏检and过检


def create_empty_sheet(test_file_path, excel_save_path):
    wb = xl.Workbook()
    file_list = os.listdir(test_file_path)
    for file in file_list:
        ws = wb.create_sheet(file, 0)  # 插入到最开始的位置(从0开始计算)
    wb.save(excel_save_path)


from openpyxl.styles import Font, Alignment


def beautify_excel(excel_path):
    font = Font(name='宋体', size=11, color='FF000000', bold=True, italic=False)
    align = Alignment(horizontal='center', vertical='center', wrap_text=False)

    book = xl.load_workbook(excel_path)  # 加载excel
    sheet_names = book.sheetnames  # 获取所有的sheet名称
    sheet_names.remove('Sheet')

    for sheet in sheet_names:
        for row in book['{}'.format(sheet)]['A1:A5']:
            for cell in row:
                cell.font = font
                cell.alignment = align

        for row in book['{}'.format(sheet)]['A12:Z19']:
            for cell in row:
                cell.alignment = align

    book.save(excel_path)


if __name__ == '__main__':
    """
    @attention：test_file_path格式！！！
    
    test_file_path
        |--- 1号面
        |--- 2号面
          。
          。
          。
        |--- 16号面
            |--- csv
            |--- gt
            |--- img
    """

    test_file_path = r"G:\weiruan_report"
    excel_save_path = r'C:\Users\Administrator\Desktop\A.xlsx'

    create_empty_sheet(test_file_path, excel_save_path)

    file_list = os.listdir(test_file_path)
    for file in file_list:
        sub_file = os.path.join(test_file_path, file)
        test_to_reports(sub_file, excel_save_path, sheet=file)

    beautify_excel(excel_save_path)
