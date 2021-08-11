"""
@Description :  
@Author      :  zhangyan 
@Time        :  2021/08/11 23:02:15 
"""

import cv2 as cv
import os

# 一寸295*413   590*826      614*860
# 一寸390*567   780*1134

def sheet1_to_name(sheet1_path, sheet_name):
    from openpyxl import load_workbook
    wb = load_workbook(sheet1_path)
    ws = wb.get_sheet_by_name(sheet_name)
    name_tuple, examinee_tuple = ws['B'], ws['L']
    name, examinee = [], []
    for i in range(1, len(name_tuple)):
        name.append(name_tuple[i].value)
        examinee.append(examinee_tuple[i].value)
    dic = dict(zip(examinee, name))
    return dic

def xywh_to_yicun(face_rec, img_w, img_h):
    x, y, w, h = face_rec[0]
    yicun_x = x - w//2
    yicun_y = y - h//1.5
    yicun_w = 2*w
    yicun_h = int(1.4*yicun_w)
    print('yicun_h:', yicun_h)
    if yicun_x < 0: yicun_x = 3
    if yicun_y < 0: yicun_y = 3
    if yicun_w > img_w: yicun_w = img_w
    if yicun_h > img_h:
        print('h过高')
        yicun_w, yicun_h = int(img_h//1.4), img_h
        yicun_x, yicun_y = x + w//2 - yicun_w//2, yicun_y
        return map(int, [yicun_x, yicun_y, yicun_w, yicun_h])
    return map(int, [yicun_x, yicun_y, yicun_w, yicun_h])


def crop_img(input_img, dic_name, output_dir):
    img = cv.imread(input_img)
    img_h, img_w = img.shape[0], img.shape[1]
    print('img_w_h:', img_w,img_h)
    classfier = cv.CascadeClassifier('/Users/zhangyan/opt/anaconda3/envs/python37/lib/python3.7/site-packages/cv2/data/haarcascade_frontalface_alt2.xml')
    face_rec = classfier.detectMultiScale(img, scaleFactor=1.2, minNeighbors=5, minSize=(50, 50))
    # print('face', face_rec)
    if len(face_rec) == 1:  # 大于0则检测到人脸
        x, y, w, h = xywh_to_yicun(face_rec, img_w, img_h)
        # print('xywh:', x, y, w, h)
        # cv.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 8)  # 3控制绿色框的粗细
        img_crop = img[y:y+h, x:x+w]
        # print('img_crop', img_crop.shape)
        img_crop = cv.resize(img_crop, (590, 826))
        padding = cv.copyMakeBorder(img_crop, 10, 24, 17, 17, cv.BORDER_CONSTANT, value=(255, 255, 255))
        img_name = os.path.basename(input_img)
        real_name = dic_name[img_name] + '.jpg'
        cv.imwrite(os.path.join(output_dir, real_name), padding)


sheet1_path = '/Users/zhangyan/Desktop/WEO_jerry/20210619第299期 溜趣/第299期_6月19日-溜趣轮滑--WE_20210614212124.xlsx'
dic_name = sheet1_to_name(sheet1_path, 'Sheet4')
# print(dic_name)

input_dir = '/Users/zhangyan/Desktop/applitcation_check'
output_dir = '/Users/zhangyan/Desktop/test'
img_list = os.listdir(input_dir)
print(img_list)
for img in img_list:
    crop_img(os.path.join(input_dir, img), dic_name, output_dir)
