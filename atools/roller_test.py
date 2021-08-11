"""
@Description:
@Author     : zhangyan
@Time       : 2021/8/11 10:30 PM
"""


def sheet1_to_name(sheet1_path, sheet_name):
    from openpyxl import load_workbook
    wb = load_workbook(sheet1_path)
    ws = wb.get_sheet_by_name(sheet_name)
    name_tuple, examinee_tuple = ws['B'], ws['L']
    name, examinee = [], []
    for i in range(1, len(name_tuple)):
        name.append(name_tuple[i].value)
        examinee.append(examinee_tuple[i].value)
    dic = dict(zip(examinee, name))
    return dic

sheet1_path = '/Users/zhangyan/Desktop/WEO_jerry/20210619第299期 溜趣/第299期_6月19日-溜趣轮滑--WE_20210614212124.xlsx'
dic_name = sheet1_to_name(sheet1_path, 'Sheet4')
# print(dic_name)
a = dic_name['序号1_6_1_mmexport1620897688325.jpg']
print(a)